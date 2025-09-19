"""
Jerusalem_water_supply.py

Script for analysis and visualization of Jerusalem water supply data.

Author: Amit Geller
Date: 2024-XX-XX
"""

import os
import tempfile
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage


def main():
    # Path to the directory containing the CSV files
    data_dir = r"C:\Users\AmitGeller\Desktop\Yaron Geller\אספקת מים לעיר ירושלים"
    
    # List all Excel files in the directory (both .xlsx and .xls), skip temp/output files
    excel_files = [
        f for f in os.listdir(data_dir)
        if (f.endswith('.xlsx') or f.endswith('.xls'))
        and not f.startswith('~$')
        and not f.startswith('water_supply_monthly_plots')
    ]
    print("Excel files found:", excel_files)
    
    # Read each Excel file into a DataFrame and store in a dict
    dfs = {}
    for file in excel_files:
        location_name = os.path.splitext(file)[0]
        df = pd.read_excel(os.path.join(data_dir, file))
        dfs[location_name] = df
        print(f"\n--- {location_name} ---")
        print("Columns:", df.columns.tolist())
        print(df.head())
    
    # Mapping for Hebrew to English location names
    hebrew_to_english = {
        'חלילים': 'Halilim',
        'חיבור חלילים': 'Halilim',
        'מוצא': 'Moza',
        'עין כרם': 'Ein cerem',
        # Add more mappings as needed
    }

    plt.figure(figsize=(16, 6))
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c']

    for i, (location, df) in enumerate(dfs.items()):
        # Map location name to English if possible
        english_location = location
        for heb, eng in hebrew_to_english.items():
            if heb in location:
                english_location = eng
                break
        if 'Date' in df.columns and 'Time' in df.columns and 'Flow' in df.columns:
            df['DateTime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str), errors='coerce')
            df_sorted = df.sort_values('DateTime')
            df_sorted['Flow_smooth'] = df_sorted['Flow'].rolling(window=10, min_periods=1).mean()
            plt.plot(df_sorted['DateTime'], df_sorted['Flow_smooth'], label=english_location, color=colors[i % len(colors)], linewidth=1)
        else:
            print(f"Required columns not found in {location}: {df.columns.tolist()}")

    plt.title('Water Flow at Jerusalem Supply Locations', fontsize=18, pad=20)
    plt.xlabel('DateTime', fontsize=14)
    plt.ylabel('Flow', fontsize=14)
    plt.legend(fontsize=12, frameon=False)
    plt.grid(True, which='major', axis='both', linestyle='-', linewidth=0.5, alpha=0.3)

    # Remove top and right spines for a cleaner look
    ax = plt.gca()
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    plt.tight_layout()
    plt.savefig(r"C:\Users\AmitGeller\Desktop\Yaron Geller\אספקת מים לעיר ירושלים\water_supply_plot.png", dpi=300)
    # plt.show()

    # Directory to save images and Excel
    save_dir = r"C:\Users\AmitGeller\Desktop\Yaron Geller\אספקת מים לעיר ירושלים"
    excel_path = os.path.join(save_dir, "water_supply_monthly_plots.xlsx")
    
    # Prepare a dict to hold monthly images
    month_images = {}
    
    # Combine all DateTime and Flow data for each location
    for_month = {}
    for location, df in dfs.items():
        if 'Date' in df.columns and 'Time' in df.columns and 'Flow' in df.columns:
            df['DateTime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str), errors='coerce')
            df_sorted = df.sort_values('DateTime')
            df_sorted['Flow_smooth'] = df_sorted['Flow'].rolling(window=10, min_periods=1).mean()
            df_sorted['Month'] = df_sorted['DateTime'].dt.to_period('M')
            for month, group in df_sorted.groupby('Month'):
                if month not in for_month:
                    for_month[month] = {}
                for_month[month][location] = group
    
    # For each month, plot all locations
    for month, loc_dict in for_month.items():
        plt.figure(figsize=(16, 6))
        for i, (location, df) in enumerate(loc_dict.items()):
            # Map location name to English if possible
            english_location = location
            for heb, eng in hebrew_to_english.items():
                if heb in location:
                    english_location = eng
                    break
            plt.plot(df['DateTime'], df['Flow_smooth'], label=english_location, color=colors[i % len(colors)], linewidth=1)
        plt.title(f'Water Flow at Jerusalem Supply Locations - {month}', fontsize=18, pad=20)
        plt.xlabel('DateTime', fontsize=14)
        plt.ylabel('Flow', fontsize=14)
        plt.legend(fontsize=12, frameon=False)
        plt.grid(True, which='major', axis='both', linestyle='-', linewidth=0.5, alpha=0.3)
        ax = plt.gca()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.tight_layout()
        img_path = os.path.join(save_dir, f'water_supply_{month}.png')
        plt.savefig(img_path, dpi=300)
        plt.close()
        month_images[str(month)] = img_path
    
    # Save all images to Excel
    wb = Workbook()
    for i, (month, img_path) in enumerate(month_images.items()):
        if i == 0:
            ws = wb.active
            ws.title = str(month)
        else:
            ws = wb.create_sheet(str(month))
        img = XLImage(img_path)
        img.width = 900
        img.height = 400
        ws.add_image(img, 'A1')
    wb.save(excel_path)
    print(f"Saved all monthly graphs to {excel_path}")

    # Individual plots for each location (full time series)
    individual_images = {}
    for i, (location, df) in enumerate(dfs.items()):
        # Map location name to English if possible
        english_location = location
        for heb, eng in hebrew_to_english.items():
            if heb in location:
                english_location = eng
                break
        if 'Date' in df.columns and 'Time' in df.columns and 'Flow' in df.columns:
            df['DateTime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str), errors='coerce')
            df_sorted = df.sort_values('DateTime')
            df_sorted['Flow_smooth'] = df_sorted['Flow'].rolling(window=10, min_periods=1).mean()
            plt.figure(figsize=(16, 6))
            plt.plot(df_sorted['DateTime'], df_sorted['Flow_smooth'], label=english_location, color=colors[i % len(colors)], linewidth=1)
            plt.title(f'Water Flow at {english_location}', fontsize=18, pad=20)
            plt.xlabel('DateTime', fontsize=14)
            plt.ylabel('Flow', fontsize=14)
            plt.legend(fontsize=12, frameon=False)
            plt.grid(True, which='major', axis='both', linestyle='-', linewidth=0.5, alpha=0.3)
            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            plt.tight_layout()
            img_path = os.path.join(save_dir, f'water_supply_{english_location}.png')
            plt.savefig(img_path, dpi=300)
            plt.close()
            individual_images[english_location] = img_path
        else:
            print(f"Required columns not found in {location}: {df.columns.tolist()}")

    # Add individual plots to Excel
    wb = load_workbook(excel_path)
    for location, img_path in individual_images.items():
        ws = wb.create_sheet(location)
        img = XLImage(img_path)
        img.width = 900
        img.height = 400
        ws.add_image(img, 'A1')
    wb.save(excel_path)
    print(f"Added individual location plots to {excel_path}")

    # --- New: Plot 4 consecutive Thursday 00:00 to Saturday 00:00 (48 hours) windows ---
    # Gather all unique Thursdays in the data
    all_datetimes = []
    for df in dfs.values():
        if 'Date' in df.columns and 'Time' in df.columns:
            dt = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str), errors='coerce')
            all_datetimes.extend(dt.dropna().tolist())
    all_datetimes = sorted(set(all_datetimes))
    thursdays = [dt for dt in all_datetimes if dt.weekday() == 3 and dt.hour == 0 and dt.minute == 0]
    thursdays = sorted(set(thursdays))
    # Only keep unique Thursdays
    thursdays = [dt for i, dt in enumerate(thursdays) if i == 0 or (dt - thursdays[i-1]).days >= 7]
    # Take up to 4 windows
    thursdays = thursdays[:4]
    thurs_sat_imgs = []
    for idx, thursday_start in enumerate(thursdays):
        saturday_start = thursday_start + pd.Timedelta(days=2)
        plt.figure(figsize=(16, 6))
        for i, (location, df) in enumerate(dfs.items()):
            # Map location name to English if possible
            english_location = location
            for heb, eng in hebrew_to_english.items():
                if heb in location:
                    english_location = eng
                    break
            if 'Date' in df.columns and 'Time' in df.columns and 'Flow' in df.columns:
                df['DateTime'] = pd.to_datetime(df['Date'].astype(str) + ' ' + df['Time'].astype(str), errors='coerce')
                mask = (df['DateTime'] >= thursday_start) & (df['DateTime'] < saturday_start)
                df_window = df.loc[mask].sort_values('DateTime')
                if not df_window.empty:
                    df_window['Flow_smooth'] = df_window['Flow'].rolling(window=10, min_periods=1).mean()
                    plt.plot(df_window['DateTime'], df_window['Flow_smooth'], label=english_location, color=colors[i % len(colors)], linewidth=1)
        title_str = f'Water Flow ({thursday_start.strftime("%Y-%m-%d %H:%M")} to {saturday_start.strftime("%Y-%m-%d %H:%M")})'
        plt.title(title_str, fontsize=18, pad=20)
        plt.xlabel('DateTime', fontsize=14)
        plt.ylabel('Flow', fontsize=14)
        plt.legend(fontsize=12, frameon=False)
        plt.grid(True, which='major', axis='both', linestyle='-', linewidth=0.5, alpha=0.3)
        ax = plt.gca()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        plt.tight_layout()
        # Save to a temporary file for Excel insertion, then delete after
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmpfile:
            plt.savefig(tmpfile.name, dpi=300)
            thurs_sat_imgs.append(tmpfile.name)
        plt.close()
    # Add all 4 graphs to a new sheet in the Excel file
    wb = load_workbook(excel_path)
    ws = wb.create_sheet('Thurs-Sat Windows')
    row = 1
    for img_path in thurs_sat_imgs:
        img = XLImage(img_path)
        img.width = 900
        img.height = 400
        ws.add_image(img, f'A{row}')
        row += 22
    wb.save(excel_path)
    # Now delete temp files
    for img_path in thurs_sat_imgs:
        os.remove(img_path)
    print(f"Added 4 Thursday-Saturday window graphs to {excel_path}")


if __name__ == "__main__":
    main() 