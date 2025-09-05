import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import os

def read_data(file_paths):
    """
    Read multiple CSV files and combine them
    """
    dfs = []
    for file_path in file_paths:
        df = pd.read_csv(file_path)
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True)

def process_data(df):
    """
    Process the data:
    1. Convert date and time columns to datetime
    2. Extract day of week and time components
    3. Filter for Mondays and Saturdays only
    4. Add daytime flag
    5. Clean invalid measurements
    """
    # Convert to datetime
    df['DateTime'] = pd.to_datetime(df['Date'] + ' ' + df['Time'])
    
    # Extract components
    df['DayOfWeek'] = df['DateTime'].dt.dayofweek
    df['DayName'] = df['DateTime'].dt.day_name()
    df['Hour'] = df['DateTime'].dt.hour
    df['Month'] = df['DateTime'].dt.strftime('%Y-%m')
    
    # Filter for Mondays (0) and Saturdays (5) only
    df = df[df['DayOfWeek'].isin([0, 5])]
    
    # Add daytime flag (08:00-15:00)
    df['IsDaytime'] = df['Hour'].between(8, 14)  # 14 to include all of hour 14 (until 15:00)
    
    # Get the water level column
    level_column = '/4sewage/bialik/balik007/level (m)'
    
    # Clean invalid measurements
    df = df[df[level_column].between(0, 10)]  # Keep only reasonable water levels
    
    return df, level_column

def analyze_data(df, level_column):
    """
    Perform detailed analysis of Monday vs Saturday data
    """
    # Overall statistics by day
    stats = df.groupby('DayName')[level_column].agg([
        'count',
        'mean',
        'median',
        'std',
        'min',
        'max',
        lambda x: x.quantile(0.25).round(3),
        lambda x: x.quantile(0.75).round(3)
    ]).round(3)
    stats.columns = ['Count', 'Mean', 'Median', 'Std', 'Min', 'Max', '25th Percentile', '75th Percentile']
    
    # Daytime statistics
    daytime_stats = df[df['IsDaytime']].groupby('DayName')[level_column].agg([
        'count',
        'mean',
        'median',
        'std',
        'min',
        'max'
    ]).round(3)
    daytime_stats.columns = ['Count', 'Mean', 'Median', 'Std', 'Min', 'Max']
    
    # Monthly statistics
    monthly_stats = df.groupby(['Month', 'DayName'])[level_column].agg([
        'count',
        'mean',
        'std'
    ]).round(3)
    
    # Hourly patterns
    hourly_stats = df.groupby(['DayName', 'Hour'])[level_column].agg(['mean', 'std']).round(3)
    
    return stats, daytime_stats, monthly_stats, hourly_stats

def create_visualizations(df, level_column, save_dir):
    """
    Create detailed visualizations comparing Mondays and Saturdays
    """
    plt.style.use('default')
    chart_paths = []
    
    # 1. Overall distribution comparison
    plt.figure(figsize=(12, 6))
    sns.boxplot(x='DayName', y=level_column, data=df)
    plt.title('Overall Water Level Distribution: Monday vs Saturday')
    plt.ylabel('Water Level (m)')
    path = os.path.join(save_dir, 'overall_distribution.png')
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    chart_paths.append(path)
    
    # 2. Daytime distribution comparison
    plt.figure(figsize=(12, 6))
    daytime_data = df[df['IsDaytime']]
    sns.boxplot(x='DayName', y=level_column, data=daytime_data)
    plt.title('Daytime (08:00-15:00) Water Level Distribution')
    plt.ylabel('Water Level (m)')
    path = os.path.join(save_dir, 'daytime_distribution.png')
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    chart_paths.append(path)
    
    # 3. Monthly trends
    plt.figure(figsize=(15, 7))
    monthly_means = df.groupby(['Month', 'DayName'])[level_column].mean().unstack()
    monthly_means.plot(marker='o')
    plt.title('Monthly Average Water Levels')
    plt.xlabel('Month')
    plt.ylabel('Average Water Level (m)')
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    path = os.path.join(save_dir, 'monthly_trends.png')
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    chart_paths.append(path)
    
    # 4. Hourly patterns with daytime highlight
    plt.figure(figsize=(15, 7))
    for day in ['Monday', 'Saturday']:
        day_data = df[df['DayName'] == day]
        hourly_mean = day_data.groupby('Hour')[level_column].mean()
        hourly_std = day_data.groupby('Hour')[level_column].std()
        plt.plot(hourly_mean.index, hourly_mean.values, label=day, marker='o')
        plt.fill_between(hourly_mean.index, 
                        hourly_mean.values - hourly_std.values,
                        hourly_mean.values + hourly_std.values,
                        alpha=0.2)
    
    # Add daytime highlight
    plt.axvspan(8, 15, color='yellow', alpha=0.2, label='Daytime (08:00-15:00)')
    plt.title('Average Water Level by Hour')
    plt.xlabel('Hour of Day')
    plt.ylabel('Water Level (m)')
    plt.legend()
    plt.grid(True, alpha=0.3)
    path = os.path.join(save_dir, 'hourly_patterns.png')
    plt.savefig(path, dpi=300, bbox_inches='tight')
    plt.close()
    chart_paths.append(path)
    
    # 5. Monthly line graphs for each month (means)
    months = sorted(df['Month'].unique())
    for month in months:
        month_data = df[df['Month'] == month].copy()
        month_data['DayOfMonth'] = month_data['DateTime'].dt.day
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 7))
        # Monday means
        monday_data = month_data[month_data['DayName'] == 'Monday']
        monday_daily = monday_data.groupby('DayOfMonth')[level_column].agg(['mean', 'std']).reset_index()
        ax1.plot(monday_daily['DayOfMonth'], monday_daily['mean'], 'b-', label='Average', linewidth=2)
        ax1.fill_between(monday_daily['DayOfMonth'],
                        monday_daily['mean'] - monday_daily['std'],
                        monday_daily['mean'] + monday_daily['std'],
                        alpha=0.2, color='blue')
        ax1.set_title(f'Monday Water Levels - {month}')
        ax1.set_xlabel('Day of Month')
        ax1.set_ylabel('Water Level (m)')
        ax1.grid(True, alpha=0.3)
        ax1.legend()
        # Saturday means
        saturday_data = month_data[month_data['DayName'] == 'Saturday']
        saturday_daily = saturday_data.groupby('DayOfMonth')[level_column].agg(['mean', 'std']).reset_index()
        ax2.plot(saturday_daily['DayOfMonth'], saturday_daily['mean'], 'r-', label='Average', linewidth=2)
        ax2.fill_between(saturday_daily['DayOfMonth'],
                        saturday_daily['mean'] - saturday_daily['std'],
                        saturday_daily['mean'] + saturday_daily['std'],
                        alpha=0.2, color='red')
        ax2.set_title(f'Saturday Water Levels - {month}')
        ax2.set_xlabel('Day of Month')
        ax2.set_ylabel('Water Level (m)')
        ax2.grid(True, alpha=0.3)
        ax2.legend()
        # Set same y-axis for both
        all_vals = pd.concat([monday_data[level_column], saturday_data[level_column]])
        if not all_vals.empty:
            y_min = all_vals.min()
            y_max = all_vals.max()
            ax1.set_ylim(y_min, y_max)
            ax2.set_ylim(y_min, y_max)
        plt.tight_layout()
        path = os.path.join(save_dir, f'monthly_line_graphs_{month}.png')
        plt.savefig(path, dpi=300, bbox_inches='tight')
        plt.close()
        chart_paths.append(path)
        # 6. Actual values for Mondays (time series)
        if not monday_data.empty:
            plt.figure(figsize=(18, 6))
            plt.plot(monday_data['DateTime'], monday_data[level_column], color='blue', marker='.', linestyle='-', linewidth=1, markersize=2)
            plt.title(f'All Monday Water Level Values - {month}')
            plt.xlabel('DateTime')
            plt.ylabel('Water Level (m)')
            plt.grid(True, alpha=0.3)
            # Set same y-axis as Saturday for this month
            if not saturday_data.empty:
                y_min = min(monday_data[level_column].min(), saturday_data[level_column].min())
                y_max = max(monday_data[level_column].max(), saturday_data[level_column].max())
                plt.ylim(y_min, y_max)
            path = os.path.join(save_dir, f'monday_actual_values_{month}.png')
            plt.savefig(path, dpi=300, bbox_inches='tight')
            plt.close()
            chart_paths.append(path)
        # 7. Actual values for Saturdays (time series)
        if not saturday_data.empty:
            plt.figure(figsize=(18, 6))
            plt.plot(saturday_data['DateTime'], saturday_data[level_column], color='red', marker='.', linestyle='-', linewidth=1, markersize=2)
            plt.title(f'All Saturday Water Level Values - {month}')
            plt.xlabel('DateTime')
            plt.ylabel('Water Level (m)')
            plt.grid(True, alpha=0.3)
            # Set same y-axis as Monday for this month
            if not monday_data.empty:
                y_min = min(monday_data[level_column].min(), saturday_data[level_column].min())
                y_max = max(monday_data[level_column].max(), saturday_data[level_column].max())
                plt.ylim(y_min, y_max)
            path = os.path.join(save_dir, f'saturday_actual_values_{month}.png')
            plt.savefig(path, dpi=300, bbox_inches='tight')
            plt.close()
            chart_paths.append(path)
    return chart_paths

def export_to_excel(df, stats, daytime_stats, monthly_stats, hourly_stats, level_column, chart_paths, save_dir):
    """
    Export analysis results to Excel
    """
    excel_file = os.path.join(save_dir, 'monday_vs_saturday_analysis.xlsx')
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Export statistics
        stats.to_excel(writer, sheet_name='Overall Statistics')
        daytime_stats.to_excel(writer, sheet_name='Daytime Statistics')
        monthly_stats.to_excel(writer, sheet_name='Monthly Statistics')
        hourly_stats.to_excel(writer, sheet_name='Hourly Statistics')
        
        # Export sample of raw data
        df_sample = df.sample(n=min(1000, len(df)))
        df_sample = df_sample.sort_values('DateTime')
        cols_to_export = ['DateTime', 'DayName', 'IsDaytime', level_column]
        df_sample[cols_to_export].to_excel(writer, sheet_name='Sample Data', index=False)
    
    # Add charts
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    
    workbook = load_workbook(excel_file)
    chart_sheet = workbook.create_sheet('Charts')
    
    months = sorted(df['Month'].unique())
    chart_titles = ['Overall Distribution', 'Daytime Distribution', 'Monthly Trends', 'Hourly Patterns']
    chart_titles += [f'Monthly Line Graphs - {month}' for month in months]
    chart_titles += [f'Monday Actual Values - {month}' for month in months]
    chart_titles += [f'Saturday Actual Values - {month}' for month in months]
    current_row = 1
    for i, (chart_path, title) in enumerate(zip(chart_paths, chart_titles)):
        img = Image(chart_path)
        img.width = 800
        img.height = 400
        chart_sheet.add_image(img, f'A{current_row}')
        current_row += 25
    
    workbook.save(excel_file)
    
    # Clean up temporary chart files
    for chart_path in chart_paths:
        os.remove(chart_path)
    
    return excel_file

def main():
    """
    Main function for Monday vs Saturday analysis
    """
    base_dir = r"C:\Users\AmitGeller\Desktop\Yaron Geller\הנס מולר 2"
    save_dir = base_dir
    
    # Define file paths for all 7 months with correct file names
    file_paths = [
        os.path.join(base_dir, "2024-10.csv"),
        os.path.join(base_dir, "2024-11.csv"),
        os.path.join(base_dir, "2024-12.csv"),
        os.path.join(base_dir, "2025-1.csv"),  # Changed from 2025-01.csv
        os.path.join(base_dir, "2025-2.csv"),  # Changed from 2025-02.csv
        os.path.join(base_dir, "2025-3.csv"),  # Changed from 2025-03.csv
        os.path.join(base_dir, "2025-4.csv")   # Changed from 2025-04.csv
    ]
    
    try:
        # Read data
        print("Reading data...")
        df = read_data(file_paths)
        
        # Process data
        print("Processing data...")
        df, level_column = process_data(df)
        
        # Analyze data
        print("Analyzing data...")
        stats, daytime_stats, monthly_stats, hourly_stats = analyze_data(df, level_column)
        
        # Create visualizations
        print("Creating visualizations...")
        chart_paths = create_visualizations(df, level_column, save_dir)
        
        # Export to Excel
        print("Exporting to Excel...")
        excel_file = export_to_excel(df, stats, daytime_stats, monthly_stats, hourly_stats, 
                                   level_column, chart_paths, save_dir)
        
        # Print summary
        print("\nAnalysis complete!")
        print(f"Results have been saved to: {excel_file}")
        print("\nOverall Statistics:")
        print(stats)
        print("\nDaytime (08:00-15:00) Statistics:")
        print(daytime_stats)
        
        # Calculate and print percentage differences
        mon_mean = stats.loc['Monday', 'Mean']
        sat_mean = stats.loc['Saturday', 'Mean']
        overall_pct_diff = ((sat_mean - mon_mean) / mon_mean) * 100
        
        mon_day_mean = daytime_stats.loc['Monday', 'Mean']
        sat_day_mean = daytime_stats.loc['Saturday', 'Mean']
        daytime_pct_diff = ((sat_day_mean - mon_day_mean) / mon_day_mean) * 100
        
        print(f"\nOverall: Water levels on Saturdays are {abs(overall_pct_diff):.1f}% {'higher' if overall_pct_diff > 0 else 'lower'} than Mondays")
        print(f"During daytime: Water levels on Saturdays are {abs(daytime_pct_diff):.1f}% {'higher' if daytime_pct_diff > 0 else 'lower'} than Mondays")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main() 